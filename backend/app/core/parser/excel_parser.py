"""openpyxl Excel 解析器,按 sheet yield 行。"""
from __future__ import annotations
from openpyxl import load_workbook
from .base import BaseParser, Chunk
from typing import Iterator


class ExcelParser(BaseParser):
    extensions = [".xlsx", ".xls"]

    def parse(self, path: str, progress_cb=None) -> Iterator[Chunk]:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        for si, name in enumerate(sheets):
            ws = wb[name]
            rows_text = []
            offset = 0
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                vals = [str(c) for c in row if c is not None]
                line = " | ".join(vals)
                if not line.strip():
                    continue
                rows_text.append(line)
                offset += len(line) + 1
            if rows_text:
                yield Chunk(
                    content="\n".join(rows_text),
                    metadata={"sheet": name, "sheet_index": si, "total_sheets": len(sheets), "rows": len(rows_text)},
                )
            if progress_cb:
                progress_cb(int((si + 1) / len(sheets) * 100))
        wb.close()